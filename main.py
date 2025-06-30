import logging
from typing import List, Optional
from playwright.sync_api import sync_playwright, Page
from dataclasses import dataclass, asdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd

import platform
import time
import datetime
import os
import re
from tqdm import tqdm
import requests
from bs4 import BeautifulSoup

@dataclass
class Place:
    name: str = ""
    maps_url: str = ""
    email: str = ""
    website: str = ""
    address: str = ""
    phone_number: str = ""
    reviews_count: Optional[int] = None
    reviews_average: Optional[float] = None
    place_type: str = ""
    introduction: str = ""
    social_media_urls: str = ""
    category: str = ""
    latitude: str = ""
    longitude: str = ""
    weekly_hours: str = ""
    tags: str = ""

def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
    )

def extract_text(page: Page, xpath: str) -> str:
    try:
        if page.locator(xpath).count() > 0:
            return page.locator(xpath).inner_text()
    except Exception as e:
        logging.warning(f"Failed to extract text for xpath {xpath}: {e}")
    return ""

def extract_email_from_website(website_url, proxies=None):
    if not website_url or website_url in ["No website", "No info available"]:
        return "No info available"
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        resp = requests.get(website_url, headers=headers, timeout=5, proxies=proxies)
        if resp.status_code == 200:
            emails = re.findall(r"[\w\.-]+@[\w\.-]+", resp.text)
            if emails:
                return ', '.join(set(emails))
    except Exception:
        pass
    return "No info available"

def extract_place(page: Page, proxies=None) -> Place:
    # XPaths (Existing)
    name_xpath = '//div[@class="TIHn2 "]//h1[@class="DUwDvf lfPIob"]'
    address_xpath = '//button[@data-item-id="address"]//div[contains(@class, "fontBodyMedium")]'
    website_xpath = '//a[@data-item-id="authority"]//div[contains(@class, "fontBodyMedium")]'
    phone_number_xpath = '//button[contains(@data-item-id, "phone:tel:")]//div[contains(@class, "fontBodyMedium")]'
    reviews_count_xpath = '//div[@class="TIHn2 "]//div[@class="fontBodyMedium dmRWX"]//div//span//span//span[@aria-label]'
    reviews_average_xpath = '//div[@class="TIHn2 "]//div[@class="fontBodyMedium dmRWX"]//div//span[@aria-hidden]'
    info1 = '//div[@class="LTs0Rc"][1]'
    info2 = '//div[@class="LTs0Rc"][2]'
    info3 = '//div[@class="LTs0Rc"][3]'
    opens_at_xpath = '//button[contains(@data-item-id, "oh")]//div[contains(@class, "fontBodyMedium")]'
    opens_at_xpath2 = '//div[@class="MkV9"]//span[@class="ZDu9vd"]//span[2]'
    place_type_xpath = '//div[@class="LBgpqf"]//button[@class="DkEaL "]'
    intro_xpath = '//div[@class="WeS02d fontBodyMedium"]//div[@class="PYvSYb "]'

    place = Place()
    place.name = extract_text(page, name_xpath) or "No info available"
    # Google Maps link extraction (use normal/long URL only)
    try:
        # Find the first place link in the left panel
        link_el = page.locator('//a[contains(@href, "/maps/place/") and not(contains(@href, "search"))]').first
        if link_el.count() > 0:
            href = link_el.get_attribute('href')
            if href and href.startswith('http'):
                place.maps_url = href
            else:
                place.maps_url = f"https://www.google.com{href}" if href else "No info available"
        else:
            place.maps_url = page.url
    except Exception:
        place.maps_url = page.url
    # Email extraction from Google Maps (if shown)
    email = "No info available"
    try:
        email_links = page.locator('//a[starts-with(@href, "mailto:")]').all()
        if email_links:
            emails = [l.get_attribute('href').replace('mailto:', '') for l in email_links if l.get_attribute('href')]
            if emails:
                email = ', '.join(set(emails))
    except Exception:
        pass
    # Email extraction from website
    website_url = extract_text(page, website_xpath)
    email_from_site = extract_email_from_website(website_url, proxies=proxies)
    if email == "No info available" and email_from_site != "No info available":
        email = email_from_site
    place.email = email
    # Website extraction
    extracted_website_url = website_url
    found_social_media_links = []
    social_media_domains_list = [
        "facebook.com", "instagram.com", "linkedin.com", "twitter.com", "x.com", "youtube.com"
    ]
    is_main_website_social = False
    if extracted_website_url:
        for domain_part in social_media_domains_list:
            if domain_part in extracted_website_url.lower():
                found_social_media_links.append(extracted_website_url)
                place.website = "No website"
                is_main_website_social = True
                break
    if not is_main_website_social:
        place.website = extracted_website_url if extracted_website_url else "No info available"
    place.address = extract_text(page, address_xpath) or "No info available"
    place.phone_number = extract_text(page, phone_number_xpath) or "No info available"
    place.place_type = extract_text(page, place_type_xpath) or "No info available"
    intro_val = extract_text(page, intro_xpath)
    place.introduction = intro_val if intro_val and intro_val.strip() else "No info available"

    # Now, find all other social media links from the page
    # This selects all 'a' tags with an 'href' attribute that contains "http"
    # within the main content area (adjust XPath if needed for your specific Google Maps view)
    all_links_in_details = page.locator('div[role="main"] a[href*="http"]').all() 
    processed_links_for_social = set()
    for link_locator in all_links_in_details:
        try:
            href = link_locator.get_attribute('href')
            if href and href not in processed_links_for_social:
                processed_links_for_social.add(href)
                href_lower = href.lower()
                for domain_part in social_media_domains_list:
                    if domain_part in href_lower:
                        found_social_media_links.append(href)
                        break
        except Exception as e:
            logging.debug(f"Failed to process a link in detail pane for social media: {e}")
    # Only keep URLs, no names
    place.social_media_urls = ", ".join(sorted(set(found_social_media_links))) or "No info available"

    # Reviews Count
    reviews_count_raw = extract_text(page, reviews_count_xpath)
    if reviews_count_raw:
        try:
            temp = reviews_count_raw.replace('\xa0', '').replace('\u202f', '').replace('(', '').replace(')', '').replace(',', '').replace(' ', '')
            place.reviews_count = int(temp)
        except Exception as e:
            logging.warning(f"Failed to parse reviews count: {e} | Raw: {reviews_count_raw} | Cleaned: {temp}")
    else:
        place.reviews_count = "No info available"
    # Reviews Average
    reviews_avg_raw = extract_text(page, reviews_average_xpath)
    if reviews_avg_raw:
        try:
            temp = reviews_avg_raw.replace(' ','').replace(',','.')
            place.reviews_average = float(temp)
        except Exception as e:
            logging.warning(f"Failed to parse reviews average: {e}")
    else:
        place.reviews_average = "No info available"
    # Remove store fields (no longer extracted)
    # Worktime: Always try to show both open and close times if available
    worktime = "No info available"
    opens_at_raw = extract_text(page, opens_at_xpath)
    open_time = close_time = None
    if opens_at_raw:
        parts = [p.strip() for p in opens_at_raw.split('⋅') if p.strip()]
        for part in parts:
            if part.lower().startswith("open"):
                open_time = part
            elif part.lower().startswith("closes") or part.lower().startswith("close"):
                close_time = part
            elif ":" in part:
                if not close_time:
                    close_time = f"Closes {part}"
        if open_time and close_time:
            worktime = f"{open_time}, {close_time}"
        elif open_time:
            worktime = open_time
        elif close_time:
            worktime = close_time
        else:
            worktime = opens_at_raw.replace("\u202f", "")
    else:
        opens_at2_raw = extract_text(page, opens_at_xpath2)
        if opens_at2_raw:
            parts = [p.strip() for p in opens_at2_raw.split('⋅') if p.strip()]
            open_time = close_time = None
            for part in parts:
                if part.lower().startswith("open"):
                    open_time = part
                elif part.lower().startswith("closes") or part.lower().startswith("close"):
                    close_time = part
                elif ":" in part:
                    if not close_time:
                        close_time = f"Closes {part}"
            if open_time and close_time:
                worktime = f"{open_time}, {close_time}"
            elif open_time:
                worktime = open_time
            elif close_time:
                worktime = close_time
            else:
                worktime = opens_at2_raw.replace("\u202f", "")
        else:
            worktime = "No info available"
    place.worktime = worktime

    # Geolocation
    lat, lng = None, None
    try:
        match = re.search(r'@([\d\.\-]+),([\d\.\-]+),', page.url)
        if match:
            lat, lng = match.group(1), match.group(2)
    except Exception:
        pass


    # Geolocation fields
    place.latitude = lat if lat else "No info available"
    place.longitude = lng if lng else "No info available"

    # Weekly hours extraction
    weekly_hours_xpath = '//table[contains(@class, "OqCZI")]'  # Google Maps hours table
    weekly_hours = "No info available"
    try:
        if page.locator(weekly_hours_xpath).count() > 0:
            table_html = page.locator(weekly_hours_xpath).inner_html()
            soup = BeautifulSoup(table_html, 'html.parser')
            rows = soup.find_all('tr')
            hours = []
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 2:
                    day = cols[0].get_text(strip=True)
                    times = cols[1].get_text(strip=True)
                    hours.append(f"{day}: {times}")
            if hours:
                weekly_hours = "; ".join(hours)
    except Exception:
        pass
    place.weekly_hours = weekly_hours
    # Tags/categories extraction (fix: use correct selector for chips/tags)
    tags = []
    try:
        tag_els = page.locator('//button[contains(@jsaction, "pane.rating.categoryChip") or contains(@class, "DkEaL")]//span').all()
        for el in tag_els:
            tag = el.inner_text().strip()
            if tag and tag not in tags:
                tags.append(tag)
    except Exception:
        pass
    place.tags = ', '.join(tags) if tags else "No info available"
    return place

def scrape_places(search_for: str, total: int, only_with_website: bool = False, browser=None, proxies=None) -> List[Place]:
    setup_logging()
    places: List[Place] = []
    unique_keys = set()
    errors = []
    close_browser = False
    if browser is None:
        with sync_playwright() as p:
            if platform.system() == "Windows":
                browser_path = r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
                browser = p.chromium.launch(executable_path=browser_path, headless=True)
            else:
                browser = p.chromium.launch(headless=True)
            close_browser = True
            result = _scrape_places_with_browser(browser, search_for, total, proxies, errors)
            if close_browser:
                browser.close()
            if errors:
                with open("scrape_errors.log", "a", encoding="utf-8") as f:
                    for err in errors:
                        f.write(err + "\n")
            return result
    else:
        result = _scrape_places_with_browser(browser, search_for, total, proxies, errors)
        if errors:
            with open("scrape_errors.log", "a", encoding="utf-8") as f:
                for err in errors:
                    f.write(err + "\n")
        return result

def _scrape_places_with_browser(browser, search_for, total, proxies=None, errors=None, lang_code='en'):
    places: List[Place] = []
    unique_keys = set()
    page = browser.new_page()
    try:
        page.goto(f"https://www.google.com/maps?hl={lang_code}", timeout=60000)
        page.wait_for_timeout(500)
        page.locator('//input[@id="searchboxinput"]').fill(search_for)
        page.keyboard.press("Enter")
        page.wait_for_selector('//a[contains(@href, "/maps/place/")]')
        page.hover('//a[contains(@href, "/maps/place/")]')
        scroll_attempts = 0
        max_scroll_attempts = 10
        last_unique_count = 0
        listings_seen = set()
        with tqdm(total=total, desc="Scraping", ncols=80) as pbar:
            while True:
                page.mouse.wheel(0, 5000)
                page.wait_for_timeout(500)
                page.wait_for_selector('//a[contains(@href, "/maps/place/")]')
                listings = page.locator('//a[contains(@href, "/maps/place/")]').all()
                listings = [listing.locator("xpath=..") for listing in listings]
                for idx, listing in enumerate(listings):
                    try:
                        listing.click()
                        page.wait_for_selector('//div[@class="TIHn2 "]//h1[@class="DUwDvf lfPIob"]', timeout=5000)
                        time.sleep(0.5)
                        place = extract_place(page, proxies=proxies)
                        key = (place.name.strip(), place.address.strip())
                        if place.name and key not in unique_keys:
                            unique_keys.add(key)
                            places.append(place)
                            pbar.update(1)
                        if len(places) >= total:
                            break
                    except Exception as e:
                        msg = f"Failed to extract listing {idx+1}: {e}"
                        if errors is not None:
                            errors.append(msg)
                        logging.warning(msg)
                if len(places) >= total:
                    break
                if len(unique_keys) == last_unique_count:
                    scroll_attempts += 1
                else:
                    scroll_attempts = 0
                last_unique_count = len(unique_keys)
                if scroll_attempts >= max_scroll_attempts:
                    logging.info("No more new unique places found after scrolling.")
                    break
    finally:
        page.close()
    return places

def save_places_to_xlsx(places: List[Place], output_path: str = "result.xlsx", append: bool = False, lang_code='en'):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    df = pd.DataFrame([asdict(place) for place in places])
    if df.empty:
        logging.warning("No data to save. DataFrame is empty.")
        return
    columns = [f.name for f in Place.__dataclass_fields__.values()]
    # Ensure email and website are right after maps_url
    if 'email' in columns and 'website' in columns:
        columns.remove('email')
        columns.remove('website')
        idx = columns.index('maps_url') + 1
        columns[idx:idx] = ['email', 'website']
    if 'category' not in columns:
        columns.append('category')
    df = df.reindex(columns=columns, fill_value="")
    header_map = get_header_translations(lang_code)
    display_headers = [header_map.get(col, col.replace('_', ' ').title()) for col in df.columns]
    file_exists = os.path.isfile(output_path)
    # Identify which columns should be hyperlinks
    hyperlink_cols = [i for i, col in enumerate(df.columns) if col in ['maps_url', 'website', 'social_media_urls']]
    if append and file_exists and output_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        start_row = worksheet.max_row + 1
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                # Make Google Maps Link, Website, Social Media URLs clickable
                if (c_idx-1) in hyperlink_cols and value and value != 'no info available':
                    # For social_media_urls, handle multiple links
                    if df.columns[c_idx-1] == 'social_media_urls' and ',' in str(value):
                        # Only link the first URL, rest as text
                        urls = [u.strip() for u in str(value).split(',') if u.strip()]
                        cell.value = urls[0]
                        cell.hyperlink = urls[0]
                        cell.style = 'Hyperlink'
                    else:
                        cell.hyperlink = value
                        cell.style = 'Hyperlink'
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for c_idx, header in enumerate(display_headers, start=1):
            worksheet.cell(row=1, column=c_idx, value=header)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                if (c_idx-1) in hyperlink_cols and value and value != 'no info available':
                    if df.columns[c_idx-1] == 'social_media_urls' and ',' in str(value):
                        urls = [u.strip() for u in str(value).split(',') if u.strip()]
                        cell.value = urls[0]
                        cell.hyperlink = urls[0]
                        cell.style = 'Hyperlink'
                    else:
                        cell.hyperlink = value
                        cell.style = 'Hyperlink'
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, size=13)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for col_cells in worksheet.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col_letter].width = max(15, min(max_len + 4, 50))
    workbook.save(output_path)
    logging.info(f"Saved {len(df)} places to {output_path} (styled Excel)")

def save_places_to_csv(places: List[Place], output_path: str = "result.csv", append: bool = False):
    
    df = pd.DataFrame([asdict(place) for place in places])
    if not df.empty:
        all_place_fields = [f.name for f in Place.__dataclass_fields__.values()]
        if 'category' not in all_place_fields:
            all_place_fields.append('category')
        df = df.reindex(columns=all_place_fields, fill_value="")

        # Write to Excel with styling
        file_exists = os.path.isfile(output_path)
        if append and file_exists and output_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            start_row = ws.max_row + 1
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Style header row
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True, size=13)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style all cells (increase font size, add padding)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        # Adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(15, min(max_length + 4, 50))

        # Save as .xlsx
        if not output_path.endswith('.xlsx'):
            output_path = output_path.rsplit('.', 1)[0] + '.xlsx'
        wb.save(output_path)
        logging.info(f"Saved {len(df)} places to {output_path} (styled Excel)")
    else:
        logging.warning("No data to save. DataFrame is empty.")

def prompt_user_input(prompt_text, example=None, allow_multiple=False, lang_code='en'):
    # Use language-specific brackets
    brackets = {
        'en': ('(', ')'),
        'ar': ('(', ')'),
        'fr': ('(', ')'),
        'de': ('(', ')'),
    }
    left, right = brackets.get(lang_code, ('(', ')'))
    example_str = f" {left}{example}{right}" if example else ""
    if allow_multiple:
        user_input = input(f"{prompt_text}{example_str} {left}{'separate multiple with commas' if lang_code == 'en' else 'افصل بين العناصر بفواصل' if lang_code == 'ar' else 'séparez par des virgules' if lang_code == 'fr' else 'durch Kommas trennen'}{right}: ")
        return [item.strip() for item in user_input.split(',') if item.strip()]
    else:
        return input(f"{prompt_text}{example_str}: ").strip()

def clean_intro_text(text):
    if not text or text.strip().lower() in ["none found", "none"]:
        return ""
    # Remove excessive whitespace, add line breaks after periods, trim
    import re
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'(?<=[.!?]) +', '\n', text)
    return text

def get_language_settings():
    languages = {
        '1': ('English', 'en'),
        '2': ('Arabic', 'ar'),
        '3': ('French', 'fr'),
        '4': ('Deutsch', 'de'),
    }
    print("\nChoose language:")
    for k, v in languages.items():
        if v[1] == 'ar':
            print(f"  {k}. {v[0]}  [!] Arabic text may look disconnected or reversed in most Windows terminals. For best results, use Windows Terminal or a GUI.")
        else:
            print(f"  {k}. {v[0]}")
    lang_choice = input("Enter 1, 2, 3, or 4: ").strip()
    if lang_choice not in languages:
        print("Invalid choice. Defaulting to English.")
        lang_choice = '1'
    lang_name, lang_code = languages[lang_choice]
    return lang_name, lang_code

def get_header_translations(lang_code):
    headers = {
        'en': {
            'name': 'Name',
            'Maps Url': 'Google Maps Link',
            'address': 'Address',
            'website': 'Website',
            'phone_number': 'Phone Number',
            'Reviews Count': 'Reviews Count',
            'Reviews Average': 'Reviews Average',
            'Place Type': 'Place Type',
            'Work Time': 'Work Time',
            'Introduction': 'Introduction',
            'Social Media': 'Social Media',
            'Category': 'Category',
            'Email': 'Email',
            'Latitude': 'Latitude',
            'Longitude': 'Longitude',
            'Weekly Hours': 'Weekly Hours',
            'Tags': 'Tags',
        },
        'ar': {
            'name': 'الاسم',
            'maps_url': 'رابط خرائط جوجل',
            'address': 'العنوان',
            'website': 'الموقع الإلكتروني',
            'phone_number': 'رقم الهاتف',
            'Reviews Count': 'عدد التقييمات',
            'Reviews Average': 'متوسط التقييمات',
            'Place Type': 'نوع المكان',
            'Work Time': 'ساعات العمل',
            'Introduction': 'مقدمة',
            'Social Media': 'وسائل التواصل',
            'Category': 'الفئة',
            'Email': 'البريد الإلكتروني',
            'Latitude': 'خط العرض',
            'Longitude': 'خط الطول',
            'Weekly Hours': 'ساعات العمل الأسبوعية',
            'Tags': 'العلامات',
        },
        'fr': {
            'name': 'Nom',
            'maps_url': 'Lien Google Maps',
            'address': 'Adresse',
            'website': 'Site Web',
            'phone_number': 'Téléphone',
            'Reviews Count': 'Nombre d’avis',
            'Reviews Average': 'Moyenne des avis',
            'Place Type': 'Type de lieu',
            'Work Time': 'Heures d’ouverture',
            'Introduction': 'Introduction',
            'Social Media': 'Réseaux sociaux',
            'Category': 'Catégorie',
            'Email': 'E-mail',
            'Latitude': 'Latitude',
            'Longitude': 'Longitude',
            'Weekly Hours': 'Horaires hebdomadaires',
            'tags': 'Étiquettes',
        },
        'de': {
            'name': 'Name',
            'maps_url': 'Google Maps Link',
            'address': 'Adresse',
            'website': 'Webseite',
            'phone_number': 'Telefonnummer',
            'Reviews Count': 'Anzahl Bewertungen',
            'Reviews Average': 'Durchschnittliche Bewertung',
            'Place Type': 'Ortstyp',
            'Work Time': 'Öffnungszeiten',
            'Introduction': 'Einführung',
            'Social Media': 'Soziale Medien',
            'category': 'Kategorie',
            'email': 'E-Mail',
            'latitude': 'Breitengrad',
            'longitude': 'Längengrad',
            'weekly_hours': 'Wöchentliche Öffnungszeiten',
            'tags': 'Tags',
        },
    }
    return headers.get(lang_code, headers['en'])

def main():
    lang_name, lang_code = get_language_settings()
    print(f"\nLanguage set to: {lang_name}\n")
    # Prompt for categories/search terms
    categories = prompt_user_input(
        "Enter categories or search terms" if lang_code == 'en' else "أدخل الفئات أو مصطلحات البحث" if lang_code == 'ar' else "Entrez les catégories ou les termes de recherche" if lang_code == 'fr' else "Geben Sie Kategorien oder Suchbegriffe ein",
        example="fine dining, casual dining, fast food" if lang_code == 'en' else "مطعم فاخر، مطعم عادي، وجبات سريعة" if lang_code == 'ar' else "gastronomie, restauration rapide" if lang_code == 'fr' else "Feinschmecker, Fast Food",
        allow_multiple=True,
        lang_code=lang_code
    )
    if not categories:
        print("No categories entered. Exiting." if lang_code == 'en' else "لم يتم إدخال أي فئات. الخروج." if lang_code == 'ar' else "Aucune catégorie saisie. Sortie." if lang_code == 'fr' else "Keine Kategorien eingegeben. Beenden.")
        return
    # Ask user if they want to search by country or city
    print("\nWould you like to search by:" if lang_code == 'en' else "هل ترغب في البحث حسب:" if lang_code == 'ar' else "Voulez-vous rechercher par :" if lang_code == 'fr' else "Möchten Sie suchen nach:")
    print("  1. Country\n  2. City" if lang_code == 'en' else "  1. دولة\n  2. مدينة" if lang_code == 'ar' else "  1. Pays\n  2. Ville" if lang_code == 'fr' else "  1. Land\n  2. Stadt")
    location_mode = input("Enter 1 for country or 2 for city: " if lang_code == 'en' else "أدخل 1 للدولة أو 2 للمدينة: " if lang_code == 'ar' else "Entrez 1 pour le pays ou 2 pour la ville :" if lang_code == 'fr' else "Geben Sie 1 für Land oder 2 für Stadt ein: ").strip()
    if location_mode == '1':
        locations = prompt_user_input("Enter country or countries" if lang_code == 'en' else "أدخل الدولة أو الدول" if lang_code == 'ar' else "Entrez le(s) pays" if lang_code == 'fr' else "Geben Sie Land/Länder ein", example="Morocco, Canada", allow_multiple=True, lang_code=lang_code)
        location_type = 'country'
    elif location_mode == '2':
        locations = prompt_user_input("Enter city or cities" if lang_code == 'en' else "أدخل المدينة أو المدن" if lang_code == 'ar' else "Entrez la/les ville(s)" if lang_code == 'fr' else "Geben Sie Stadt/Städte ein", example="kenitra, rabat", allow_multiple=True, lang_code=lang_code)
        location_type = 'city'
    else:
        print("Invalid choice. Exiting." if lang_code == 'en' else "خيار غير صالح. الخروج." if lang_code == 'ar' else "Choix invalide. Sortie." if lang_code == 'fr' else "Ungültige Auswahl. Beenden.")
        return
    if not locations:
        print((f"No {location_type} entered. Exiting." if lang_code == 'en' else f"لم يتم إدخال أي {location_type}. الخروج." if lang_code == 'ar' else f"Aucun(e) {location_type} saisi(e). Sortie." if lang_code == 'fr' else f"Keine {location_type} eingegeben. Beenden."))
        return
    # Prompt for export type
    print("\n" + ("Choose export type:" if lang_code == 'en' else "اختر نوع التصدير:" if lang_code == 'ar' else "Choisissez le type d'exportation :" if lang_code == 'fr' else "Exporttyp wählen:"))
    print("  1. Excel\n  2. JSON\n  3. HTML" if lang_code == 'en' else "  1. إكسل\n  2. جيسون\n  3. إتش تي إم إل" if lang_code == 'ar' else "  1. Excel\n  2. JSON\n  3. HTML" if lang_code == 'fr' else "  1. Excel\n  2. JSON\n  3. HTML")
    export_choice = input(("Enter 1, 2, or 3: " if lang_code == 'en' else "أدخل 1 أو 2 أو 3: " if lang_code == 'ar' else "Entrez 1, 2 ou 3 : " if lang_code == 'fr' else "Geben Sie 1, 2 oder 3 ein: ")).strip()
    export_type = None
    if export_choice == '1':
        export_type = 'excel'
    elif export_choice == '2':
        export_type = 'json'
    elif export_choice == '3':
        export_type = 'html'
    else:
        print("Invalid export type. Exiting." if lang_code == 'en' else "نوع التصدير غير صالح. الخروج." if lang_code == 'ar' else "Type d'exportation invalide. Sortie." if lang_code == 'fr' else "Ungültiger Exporttyp. Beenden.")
        return
    # Prompt for number of results
    try:
        total = int(prompt_user_input(
            "How many results per category?" if lang_code == 'en' else "كم عدد النتائج لكل فئة؟" if lang_code == 'ar' else "Combien de résultats par catégorie ?" if lang_code == 'fr' else "Wie viele Ergebnisse pro Kategorie?",
            example="10",
            lang_code=lang_code
        ))
    except Exception:
        print("Invalid number. Exiting." if lang_code == 'en' else "رقم غير صالح. الخروج." if lang_code == 'ar' else "Nombre invalide. Sortie." if lang_code == 'fr' else "Ungültige Zahl. Beenden.")
        return
    # Ask if user wants to see browser
    print("\n" + ("Do you want to see the browser while scraping?" if lang_code == 'en' else "هل تريد رؤية المتصفح أثناء الاستخراج؟" if lang_code == 'ar' else "Voulez-vous voir le navigateur pendant le scraping ?" if lang_code == 'fr' else "Möchten Sie den Browser beim Scraping sehen?"))
    print("  1. Yes (see data getting scraped)" if lang_code == 'en' else "  1. نعم (شاهد البيانات أثناء الاستخراج)" if lang_code == 'ar' else "  1. Oui (voir les données extraites)" if lang_code == 'fr' else "  1. Ja (Daten werden angezeigt)")
    print("  2. No  (run in background, faster)" if lang_code == 'en' else "  2. لا (تشغيل في الخلفية، أسرع)" if lang_code == 'ar' else "  2. Non (exécution en arrière-plan, plus rapide)" if lang_code == 'fr' else "  2. Nein (im Hintergrund, schneller)")
    headless_choice = input(("Enter 1 for visible, 2 for headless: " if lang_code == 'en' else "أدخل 1 للظهور، 2 للخلفية: " if lang_code == 'ar' else "Entrez 1 pour visible, 2 pour arrière-plan :" if lang_code == 'fr' else "Geben Sie 1 für sichtbar, 2 für Hintergrund ein: ")).strip()
    headless = False if headless_choice == '1' else True
    output_dir = "results"
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    location_results = {}
    jobs = [(cat, loc) for cat in categories for loc in locations]
    # If only one job, run as before
    if len(jobs) == 1:
        from playwright.sync_api import sync_playwright
        print("Opening the browser..." if lang_code == 'en' else "جاري فتح المتصفح..." if lang_code == 'ar' else "Ouverture du navigateur..." if lang_code == 'fr' else "Browser wird geöffnet...")
        with sync_playwright() as p:
            if platform.system() == "Windows":
                browser_path = r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
                browser = p.chromium.launch(executable_path=browser_path, headless=headless)
            else:
                browser = p.chromium.launch(headless=headless)
            cat, loc = jobs[0]
            full_query = f"{cat} in {loc}"
            places = scrape_places(full_query, total, browser=browser, proxies=None)
            for place in places:
                place.category = cat
            location_results[loc] = places
            browser.close()
    else:
        from concurrent.futures import ThreadPoolExecutor
        from playwright.sync_api import sync_playwright
        def scrape_job(args):
            cat, loc = args
            print((f"Opening the browser for: {cat} in {loc} ..." if lang_code == 'en' else f"جاري فتح المتصفح لـ: {cat} في {loc} ..." if lang_code == 'ar' else f"Ouverture du navigateur pour : {cat} à {loc} ..." if lang_code == 'fr' else f"Browser wird geöffnet für: {cat} in {loc} ..."))
            with sync_playwright() as p:
                if platform.system() == "Windows":
                    browser_path = r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
                    browser = p.chromium.launch(executable_path=browser_path, headless=headless)
                else:
                    browser = p.chromium.launch(headless=headless)
                full_query = f"{cat} in {loc}"
                places = scrape_places(full_query, total, browser=browser, proxies=None)
                for place in places:
                    place.category = cat
                browser.close()
                return (loc, places)
        with ThreadPoolExecutor(max_workers=len(jobs)) as executor:
            results = list(executor.map(scrape_job, jobs))
        for loc, places in results:
            if loc not in location_results:
                location_results[loc] = []
            location_results[loc].extend(places)
    # Export logic
    if export_type == 'excel':
        output_path = os.path.join(output_dir, f"Maps_scrape_{'_'.join([l for l in locations])}_{timestamp}.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        for location, places in location_results.items():
            ws = wb.create_sheet(title=location[:31])  # Excel sheet name max 31 chars
            df = pd.DataFrame([asdict(place) for place in places])
            columns = [f.name for f in Place.__dataclass_fields__.values()]
            # Ensure email and website are right after maps_url
            if 'email' in columns and 'website' in columns:
                columns.remove('email')
                columns.remove('website')
                idx = columns.index('maps_url') + 1
                columns[idx:idx] = ['email', 'website']
            if 'category' not in columns:
                columns.append('category')
            df = df.reindex(columns=columns, fill_value="")
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            # Style header row
            header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            header_font = Font(bold=True, size=13)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Style data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            # Auto-size columns
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(15, min(max_len + 4, 50))
        wb.save(output_path)
        print(f"\n\033[92mResults saved to {output_path}\033[0m\n")
    elif export_type == 'json':
        output_path = os.path.join(output_dir, f"Maps_scrape_{'_'.join([l for l in locations])}_{timestamp}.json")
        import json
        all_places = []
        for location, places in location_results.items():
            for place in places:
                place_dict = asdict(place)
                place_dict[location_type] = location
                all_places.append(place_dict)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(all_places, f, ensure_ascii=False, indent=2)
        print(f"\n\033[92mResults saved to {output_path}\033[0m\n")
    elif export_type == 'html':
        output_path = os.path.join(output_dir, f"Maps_scrape_{'_'.join([l for l in locations])}_{timestamp}.html")
        all_places = []
        for location, places in location_results.items():
            for place in places:
                place_dict = asdict(place)
                place_dict[location_type] = location
                all_places.append(place_dict)
        df = pd.DataFrame(all_places)
        df.to_html(output_path, index=False, border=1, justify='center')
        print(f"\n\033[92mResults saved to {output_path}\033[0m\n")
    else:
        print("Unknown export type.")

    # Show a clean summary of results
    print("\n==============================")
    print("         SUMMARY")
    print("==============================")
    for location, places in location_results.items():
        print(f"\nLocation: \033[94m{location}\033[0m | Total businesses: \033[93m{len(places)}\033[0m")
        for idx, place in enumerate(places, 1):
            print(f"  {idx}. \033[1m{place.name}\033[0m")
            print(f"     Address: {place.address}")
            print(f"     Website: {place.website}")
            print(f"     Phone: {place.phone_number}")
            print(f"     Reviews: {place.reviews_count} (Avg: {place.reviews_average})")
            print(f"     Category: {place.category}")
            print(f"     Social: {place.social_media_urls}")
            print(f"     Email: {place.email}")
            print(f"     Latitude: {place.latitude}")
            print(f"     Longitude: {place.longitude}")
            print(f"     Weekly Hours: {place.weekly_hours}")
            print(f"     Tags: {place.tags}")
            # Clean intro/info formatting
            intro = place.introduction.strip().replace('\n', ' ').replace('  ', ' ')
            if intro and intro != 'None Found':
                print(f"     Introduction: \033[90m{intro}\033[0m")
            print("     ---")
    print("\n==============================\n")

if __name__ == "__main__":
    main()